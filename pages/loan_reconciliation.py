import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill

# ─────────────────────────────────────────────
#  STYLE — match El Imperio UI
# ─────────────────────────────────────────────
st.markdown("""
<style>
    /* Hide sidebar entirely */
    [data-testid="stSidebar"] { display: none; }

    /* Header bar matching El Imperio */
    .ei-header {
        display: flex;
        align-items: center;
        border-bottom: 2px solid #1a2744;
        padding-bottom: 1rem;
        margin-bottom: 1.5rem;
    }
    .ei-header-text h1 {
        font-size: 2rem;
        font-weight: 700;
        color: #1a2744;
        margin: 0;
        letter-spacing: 0.02em;
    }
    .ei-header-text p {
        font-size: 0.75rem;
        letter-spacing: 0.15em;
        color: #888;
        margin: 0;
        text-transform: uppercase;
    }

    /* Section labels */
    .section-label {
        font-size: 0.7rem;
        letter-spacing: 0.15em;
        text-transform: uppercase;
        color: #1a2744;
        font-weight: 600;
        margin-bottom: 0.5rem;
    }

    /* Green download button */
    .stDownloadButton > button {
        background-color: #1e7e34 !important;
        color: white !important;
        font-weight: 600 !important;
        font-size: 1rem !important;
        padding: 0.65rem 2rem !important;
        border: none !important;
        border-radius: 4px !important;
        width: 100% !important;
    }
    .stDownloadButton > button:hover {
        background-color: #155724 !important;
    }

    /* Run button */
    .stButton > button[kind="primary"] {
        background-color: #1a2744 !important;
        color: white !important;
        font-weight: 600 !important;
        border-radius: 4px !important;
        width: 100% !important;
    }

    /* Metric cards */
    [data-testid="stMetric"] {
        background: #f8f9fa;
        border: 1px solid #e0e0e0;
        border-radius: 6px;
        padding: 1rem;
    }
    [data-testid="stMetricLabel"] { color: #1a2744 !important; font-weight: 600; }
    [data-testid="stMetricValue"] { color: #1a2744 !important; }
</style>
""", unsafe_allow_html=True)

# Header
st.markdown("""
<div class="ei-header">
    <div class="ei-header-text">
        <h1>Loan Reconciliation</h1>
        <p>Intercompany Loan Matching &middot; Powered by El Imperio</p>
    </div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
#  FORMAT CONFIGS
# ─────────────────────────────────────────────
FORMATS = {
    "Sage":   {"date_col": "Account Date"},
    "Pastel": {"date_col": "Date"},
}

FILL_GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────

def load_and_normalise(uploaded_file, fmt, company_label):
    try:
        raw = uploaded_file.read().decode("utf-8", errors="replace")
        uploaded_file.seek(0)
        first_line = raw.splitlines()[0].strip().lower()
        skip = 1 if first_line.startswith("sep=") else 0
        df = pd.read_csv(uploaded_file, dtype=str, skiprows=skip)
    except Exception as e:
        st.error(f"{company_label}: Could not read CSV — {e}")
        return None, None

    date_col = FORMATS[fmt]["date_col"]

    if date_col not in df.columns:
        st.error(
            f"{company_label}: Expected column '{date_col}' not found. "
            f"Available columns: {list(df.columns)}"
        )
        return None, None

    mask = df[date_col].str.strip().str.match(r"^\d{2}/\d{2}/\d{4}$", na=False)
    df = df[mask].copy().reset_index(drop=True)

    if df.empty:
        st.error(f"{company_label}: No valid transaction rows found after filtering.")
        return None, None

    export_df = df.copy()

    clean = pd.DataFrame()
    clean["date"]        = pd.to_datetime(df[date_col].str.strip(), format="%d/%m/%Y")
    clean["description"] = df["Description"].fillna("").astype(str).str.strip()
    clean["debit"]       = pd.to_numeric(df["Debit"].str.replace(",", ""), errors="coerce").fillna(0)
    clean["credit"]      = pd.to_numeric(df["Credit"].str.replace(",", ""), errors="coerce").fillna(0)

    return clean.reset_index(drop=True), export_df.reset_index(drop=True)


def reconcile(df_a, df_b, tolerance=3):
    a = df_a.copy()
    b = df_b.copy()
    a["_used"]   = False
    b["_used"]   = False
    a["_status"] = "unmatched"
    b["_status"] = "unmatched"

    confirmed = []
    uncertain = []

    for pass_num in range(2):
        for i, ra in a.iterrows():
            if a.at[i, "_used"]:
                continue

            if ra["debit"] > 0:
                amount       = ra["debit"]
                b_amount_col = "credit"
            elif ra["credit"] > 0:
                amount       = ra["credit"]
                b_amount_col = "debit"
            else:
                continue

            candidates = []
            for j, rb in b.iterrows():
                if b.at[j, "_used"]:
                    continue
                if abs(rb[b_amount_col] - amount) < 0.005:
                    day_diff = abs((ra["date"] - rb["date"]).days)
                    if pass_num == 0 and day_diff == 0:
                        candidates.append((j, day_diff))
                    elif pass_num == 1 and 0 < day_diff <= tolerance:
                        candidates.append((j, day_diff))

            if candidates:
                candidates.sort(key=lambda x: x[1])
                j, day_diff = candidates[0]
                status = "confirmed" if pass_num == 0 else "uncertain"

                a.at[i, "_used"]   = True
                b.at[j, "_used"]   = True
                a.at[i, "_status"] = status
                b.at[j, "_status"] = status

                record = {
                    "amount":        amount,
                    "a_date":        ra["date"],
                    "a_description": ra["description"],
                    "a_debit":       ra["debit"],
                    "a_credit":      ra["credit"],
                    "b_date":        b.at[j, "date"],
                    "b_description": b.at[j, "description"],
                    "b_debit":       b.at[j, "debit"],
                    "b_credit":      b.at[j, "credit"],
                    "day_diff":      day_diff,
                }
                if pass_num == 0:
                    confirmed.append(record)
                else:
                    uncertain.append(record)

    unmatched_a = a[~a["_used"]].drop(columns=["_used", "_status"]).reset_index(drop=True)
    unmatched_b = b[~b["_used"]].drop(columns=["_used", "_status"]).reset_index(drop=True)
    status_a    = a["_status"].reset_index(drop=True)
    status_b    = b["_status"].reset_index(drop=True)

    return confirmed, uncertain, unmatched_a, unmatched_b, status_a, status_b


def fmt_amount(v):
    return f"R {v:,.2f}" if v else "-"


def fmt_date(d):
    return d.strftime("%d/%m/%Y") if pd.notna(d) else "-"


def matches_to_df(matches, name_a, name_b):
    if not matches:
        return pd.DataFrame()
    rows = []
    for m in matches:
        rows.append({
            "Amount":                fmt_amount(m["amount"]),
            f"{name_a} Date":        fmt_date(m["a_date"]),
            f"{name_a} Description": m["a_description"],
            f"{name_a} Debit":       fmt_amount(m["a_debit"]),
            f"{name_a} Credit":      fmt_amount(m["a_credit"]),
            f"{name_b} Date":        fmt_date(m["b_date"]),
            f"{name_b} Description": m["b_description"],
            f"{name_b} Debit":       fmt_amount(m["b_debit"]),
            f"{name_b} Credit":      fmt_amount(m["b_credit"]),
            "Day Difference":        m["day_diff"],
        })
    return pd.DataFrame(rows)


def unmatched_to_df(df, missing_in):
    rows = []
    for _, r in df.iterrows():
        if r["debit"] > 0:
            amount     = r["debit"]
            needs_type = "Credit"
        elif r["credit"] > 0:
            amount     = r["credit"]
            needs_type = "Debit"
        else:
            continue
        rows.append({
            "Date":            fmt_date(r["date"]),
            "Description":     r["description"],
            "Debit":           fmt_amount(r["debit"]),
            "Credit":          fmt_amount(r["credit"]),
            "Action Required": f"{missing_in} needs a {needs_type} of {fmt_amount(amount)} around {fmt_date(r['date'])}",
        })
    return pd.DataFrame(rows)


def write_highlighted_sheet(writer, export_df, status_series, sheet_name):
    export_df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    for row_idx, status in enumerate(status_series, start=2):
        if status == "confirmed":
            fill = FILL_GREEN
        elif status == "uncertain":
            fill = FILL_YELLOW
        else:
            continue
        for cell in ws[row_idx]:
            cell.fill = fill


def to_excel(confirmed, uncertain, unmatched_a, unmatched_b,
             export_a, export_b, status_a, status_b, name_a, name_b):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        write_highlighted_sheet(writer, export_a, status_a, f"{name_a[:28]}")
        write_highlighted_sheet(writer, export_b, status_b, f"{name_b[:28]}")
        matches_to_df(confirmed, name_a, name_b).to_excel(writer, sheet_name="Confirmed Matches", index=False)
        matches_to_df(uncertain, name_a, name_b).to_excel(writer, sheet_name="Uncertain Matches", index=False)
        unmatched_to_df(unmatched_a, name_b).to_excel(writer, sheet_name=f"Unmatched {name_a[:20]}", index=False)
        unmatched_to_df(unmatched_b, name_a).to_excel(writer, sheet_name=f"Unmatched {name_b[:20]}", index=False)
    return buf.getvalue()


# ─────────────────────────────────────────────
#  STEP 1 — INPUTS
# ─────────────────────────────────────────────
st.markdown('<p class="section-label">Step 1 — Company Details</p>', unsafe_allow_html=True)

col_a, col_b, col_set = st.columns([2, 2, 1])

with col_a:
    st.markdown('<p class="section-label">Company A</p>', unsafe_allow_html=True)
    name_a = st.text_input("Company name", value="Company A", key="name_a", label_visibility="collapsed")
    fmt_a  = st.selectbox("Export format", list(FORMATS.keys()), key="fmt_a")
    file_a = st.file_uploader("Upload CSV", type="csv", key="file_a")

with col_b:
    st.markdown('<p class="section-label">Company B</p>', unsafe_allow_html=True)
    name_b = st.text_input("Company name", value="Company B", key="name_b", label_visibility="collapsed")
    fmt_b  = st.selectbox("Export format", list(FORMATS.keys()), key="fmt_b")
    file_b = st.file_uploader("Upload CSV", type="csv", key="file_b")

with col_set:
    st.markdown('<p class="section-label">Settings</p>', unsafe_allow_html=True)
    date_tolerance = st.slider("Date tolerance (days)", min_value=0, max_value=7, value=3)

st.divider()
run_btn = st.button("Run Reconciliation", type="primary", use_container_width=True)

# ─────────────────────────────────────────────
#  STEP 2 — RESULTS
# ─────────────────────────────────────────────
if not run_btn:
    st.info("Fill in company details, upload both CSVs above, then click Run Reconciliation.")
    st.stop()

if not file_a or not file_b:
    st.warning("Please upload a CSV for both companies before running.")
    st.stop()

with st.spinner("Loading files..."):
    df_a, export_a = load_and_normalise(file_a, fmt_a, name_a)
    df_b, export_b = load_and_normalise(file_b, fmt_b, name_b)

if df_a is None or df_b is None:
    st.stop()

with st.spinner("Matching transactions..."):
    confirmed, uncertain, unmatched_a, unmatched_b, status_a, status_b = reconcile(df_a, df_b, date_tolerance)

st.markdown('<p class="section-label">Step 2 — Results</p>', unsafe_allow_html=True)

c1, c2, c3, c4 = st.columns(4)
c1.metric("Confirmed Matches",      len(confirmed))
c2.metric("Uncertain Matches",      len(uncertain))
c3.metric(f"Unmatched in {name_a}", len(unmatched_a))
c4.metric(f"Unmatched in {name_b}", len(unmatched_b))

st.divider()

try:
    excel_bytes = to_excel(
        confirmed, uncertain, unmatched_a, unmatched_b,
        export_a, export_b, status_a, status_b, name_a, name_b
    )
    st.download_button(
        label="Download Full Report (Excel)",
        data=excel_bytes,
        file_name="loan_reconciliation.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
except Exception as e:
    st.warning(f"Excel export unavailable: {e}. Make sure openpyxl is in your requirements.txt.")

st.divider()

with st.expander(f"Confirmed Matches  ({len(confirmed)})", expanded=True):
    if confirmed:
        st.dataframe(matches_to_df(confirmed, name_a, name_b), use_container_width=True, hide_index=True)
    else:
        st.write("No confirmed matches found.")

with st.expander(f"Uncertain Matches — review these  ({len(uncertain)})", expanded=True):
    if uncertain:
        df_unc = matches_to_df(uncertain, name_a, name_b)
        st.dataframe(
            df_unc.style.map(lambda v: "background-color: #fff3cd", subset=["Day Difference"]),
            use_container_width=True,
            hide_index=True,
        )
    else:
        st.write("No uncertain matches.")

with st.expander(f"Unmatched in {name_a}  ({len(unmatched_a)})", expanded=True):
    if not unmatched_a.empty:
        df_ua = unmatched_to_df(unmatched_a, name_b)
        st.dataframe(
            df_ua.style.map(lambda v: "background-color: #f8d7da", subset=["Action Required"]),
            use_container_width=True,
            hide_index=True,
        )
    else:
        st.write(f"All {name_a} transactions matched.")

with st.expander(f"Unmatched in {name_b}  ({len(unmatched_b)})", expanded=True):
    if not unmatched_b.empty:
        df_ub = unmatched_to_df(unmatched_b, name_a)
        st.dataframe(
            df_ub.style.map(lambda v: "background-color: #f8d7da", subset=["Action Required"]),
            use_container_width=True,
            hide_index=True,
        )
    else:
        st.write(f"All {name_b} transactions matched.")
